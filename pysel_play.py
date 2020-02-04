#!/usr/bin/env python3
import os
import sys
import time
from selenium import webdriver
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.action_chains import ActionChains
import json
import openpyxl as px

# パラメタチェック
if len(sys.argv)<2 :
    print("{} webscenariofile".format(sys.argv[0]))
    exit(1)

websenariofile=sys.argv[1]

options = Options()

options.binary_location='/usr/bin/firefox'
if not os.path.isfile(options.binary_location) :
    options.binary_location='/usr/local/bin/firefox'
#options.add_argument('--headless')
    
driver = webdriver.Firefox(firefox_options=options)
driver.set_window_size(1200,600)
driver.set_window_position(0,0)

actions = ActionChains(driver)

wb=px.Workbook()

# 要素が見つかるまでの待ち時間
driver.implicitly_wait(5)


jsonfd=open(websenariofile,"r")
json  = json.load(jsonfd)

#print(json)
head= json["head"]
print("title:"+head["title"])
waitTime=int(head["waitTime"])

steps=json["step"]

for step in steps :
    print(step["description"])
    command = step["command"]
    if command == "navigate":
        driver.get(step["url"])
        print("navigate: "+step["url"])
    if command == "send_keys":
        elm = driver.find_element_by_xpath(step["elm_xpath"])
        elm.send_keys(step["send_keys"])
    if command == "submit":
        elm = driver.find_element_by_xpath(step["elm_xpath"])
        elm.submit()
    if command == "click":
        elm = driver.find_element_by_xpath(step["elm_xpath"])
        elm.click()
    if command == "move_to_element":
        elm = driver.find_element_by_xpath(step["elm_xpath"])
        actions.move_to_element(elm).perform()
    if command == "scroll_element_to":
        elm = driver.find_element_by_xpath(step["elm_xpath"])
        driver.execute_script("arguments[0].scrollTop={}".format(step["y-coord"]),elm)
        driver.execute_script("arguments[0].scrollLeft={}".format(step["x-coord"]),elm)
    if command == "getElemByXpath":
        elm = driver.find_element_by_xpath(step["elm_xpath"])
        print("取得テキスト: 「"+elm.text+"」")
    if command == "refresh":
        driver.refresh()

    time.sleep(waitTime)

    if "screenshot_file" in step:
        driver.save_screenshot(step["screenshot_file"])
        if "screenshot_xlsx" in step:
            ws=wb.create_sheet(title=step["screenshot_xlsx"])
            ws["A1"].value=step["description"]
            img = px.drawing.image.Image(step["screenshot_file"])
            ws.add_image(img,'A2')

wb.save(head["screenshot_xlsxfile"])
        
    
#driver.close()

